#!/usr/bin/env python3
"""Extract senior results from the Coupe de France de Robotique Google Sheet
into clean JSONs, for any year.

Usage: python scripts/extract_results.py <year>
       expects   data/coupe<year>/sheet.xlsx
       writes    data/coupe<year>/teams_results.json    (tab 'Résultats équipes')
                 data/coupe<year>/matches_results.json  (tab 'Matchs Agrégés')
"""

import json
import sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# Column indices (1-based). Row 2 has the headers; data starts row 3 (row 3 is
# a 'Placeholder' line in 2026, skipped via the name check).
COLS = {
    "type": 1, "name": 2, "rank": 4, "total": 5,
    "s1": 6, "s2": 7, "s3": 8, "s4": 9, "s5": 10,
    "matchs": 11, "wins": 12, "losses": 13, "draws": 14,
    "avg": 15, "std": 16, "min": 17, "max": 18,
    "huitiemes": 19, "quarts": 20, "semi": 21,
    "petite_finale": 22, "finale_1": 23, "finale_2": 24, "finale_3": 25,
}
PHASES = ("huitiemes", "quarts", "semi", "petite_finale",
          "finale_1", "finale_2", "finale_3")


def clean(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


# 'Matchs Agrégés' columns (1-based). Row 1 = section header, row 2 = sub
# headers, data from row 3. Note: 2024 uses 'EQUIPE 1/EQUIPE 2' (no
# blue/yellow info) — we store everything as side_a/side_b in that case.
M_COLS = {"n": 1, "series": 2, "name_a": 3, "score_a": 4, "winner": 5,
          "score_b": 6, "name_b": 7}
# Series cell -> normalized series key. Anything text-like (Huitièmes, Quarts,
# Finale 1…) folds into "finales" with a phase sub-field.
FINALES_PHASES = {
    "Huitièmes": "huitiemes", "Quarts": "quarts", "Semi-F.": "semi",
    "P.Finale": "petite_finale", "Finale 1": "finale_1",
    "Finale 2": "finale_2", "Finale 3": "finale_3",
    "Semi-F. Legends": "semi_legends", "Finale Legends": "finale_legends",
}


def _series_key(raw):
    """Return ('1'..'5', None) for regular series, ('finales', phase) otherwise."""
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)):
        return str(int(raw)), None
    phase = FINALES_PHASES.get(str(raw).strip())
    return ("finales", phase) if phase else (None, None)


def extract_teams(wb, year):
    ws = wb["Résultats  équipes"]            # tab name has TWO spaces
    teams = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = clean(row[COLS["name"] - 1])
        if not name or name.lower() == "placeholder":
            continue
        teams.append({
            "name": name,
            "type": "legends" if clean(row[COLS["type"] - 1]) == "Legends" else None,
            "rank": clean(row[COLS["rank"] - 1]),
            "score_total": clean(row[COLS["total"] - 1]),
            "scores": [clean(row[COLS[f"s{i}"] - 1]) for i in range(1, 6)],
            "matches": clean(row[COLS["matchs"] - 1]),
            "wins": clean(row[COLS["wins"] - 1]),
            "losses": clean(row[COLS["losses"] - 1]),
            "draws": clean(row[COLS["draws"] - 1]),
            "score_avg": clean(row[COLS["avg"] - 1]),
            "score_std": clean(row[COLS["std"] - 1]),
            "score_min": clean(row[COLS["min"] - 1]),
            "score_max": clean(row[COLS["max"] - 1]),
            "phases": {p: clean(row[COLS[p] - 1]) for p in PHASES},
        })
    return teams


def extract_matches(wb, year):
    if "Matchs Agrégés" not in wb.sheetnames:
        return []                           # 2024 sometimes has no aggregate tab
    ws = wb["Matchs Agrégés"]
    # In 2025/2026, name_a = bleu and name_b = jaune (per the EQUIPE BLEU /
    # EQUIPE JAUNE section headers on row 1). In 2024, name_a/b = 1/2 (no color
    # info) — we set the colors to None so the joiner uses set-matching.
    has_colors = ws.cell(row=1, column=3).value == "EQUIPE BLEU"
    matches = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        s_key, phase = _series_key(row[M_COLS["series"] - 1])
        if not s_key:
            continue
        name_a = clean(row[M_COLS["name_a"] - 1])
        name_b = clean(row[M_COLS["name_b"] - 1])
        if not name_a or not name_b:
            continue
        sa = clean(row[M_COLS["score_a"] - 1])
        sb = clean(row[M_COLS["score_b"] - 1])
        w = clean(row[M_COLS["winner"] - 1])
        winner = ("a" if w == 1 else "b" if w == -1 else None)
        m = {
            "series": s_key, "n": clean(row[M_COLS["n"] - 1]),
            "team_blue" if has_colors else "team_a": name_a,
            "team_yellow" if has_colors else "team_b": name_b,
            "score_blue" if has_colors else "score_a": sa,
            "score_yellow" if has_colors else "score_b": sb,
            "winner": (("blue" if winner == "a" else "yellow") if has_colors
                       else ("a" if winner == "a" else "b" if winner == "b" else None))
                      if winner else None,
            "colors_known": has_colors,
        }
        if phase:
            m["phase"] = phase
        matches.append(m)
    return matches


def extract(year):
    sheet = ROOT / f"data/coupe{year}/sheet.xlsx"
    wb = openpyxl.load_workbook(sheet, data_only=True)

    teams = extract_teams(wb, year)
    out_t = ROOT / f"data/coupe{year}/teams_results.json"
    out_t.parent.mkdir(parents=True, exist_ok=True)
    out_t.write_text(json.dumps({
        "year": year, "category": "senior", "teams": teams,
    }, ensure_ascii=False, indent=2))
    n_phases = sum(1 for t in teams if any(t["phases"][p] is not None for p in PHASES))
    n_legends = sum(1 for t in teams if t["type"] == "legends")
    print(f"[{year}] {len(teams)} équipes, {n_phases} en phases finales, "
          f"{n_legends} Legends -> {out_t.relative_to(ROOT)}")

    matches = extract_matches(wb, year)
    if matches:
        out_m = ROOT / f"data/coupe{year}/matches_results.json"
        out_m.write_text(json.dumps({
            "year": year, "category": "senior", "matches": matches,
        }, ensure_ascii=False, indent=2))
        by_s = {}
        for m in matches:
            by_s[m["series"]] = by_s.get(m["series"], 0) + 1
        print(f"[{year}] {len(matches)} matchs avec scores ({by_s}) -> "
              f"{out_m.relative_to(ROOT)}")


if __name__ == "__main__":
    extract(int(sys.argv[1]))
